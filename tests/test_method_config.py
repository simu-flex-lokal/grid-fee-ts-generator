from pathlib import Path

import pandas as pd
import pytest
from openpyxl import Workbook

from grid_fee.method_config import (
    METHOD_SHEET_NAMES,
    load_methods_config_from_excel,
)
from grid_fee.methods import create_methods_from_config

EXAMPLE_CONFIG = Path(__file__).resolve().parents[1] / "examples" / "method_config.xlsx"


def _sample_config() -> dict[str, dict[str, object]]:
    return {
        "topn_peak_reference_day": {
            "base_fee": 10.0,
            "relative_fee_reduction": 0.4,
            "relative_fee_surcharge": 0.2,
            "n_low_peaks": 1,
            "n_high_peaks": 3,
            "window_size_hours_low": 2.0,
            "window_size_hours_high": 6.0,
            "use_reference_day": False,
        },
        "quantile_daily_budget": {
            "base_fee": 10.0,
            "relative_fee_reduction": 0.3,
            "relative_fee_surcharge": 0.2,
            "q_low": 0.2,
            "q_high": 0.1,
            "selection_mode": "contiguous",
            "max_blocks_low": 2,
            "max_blocks_high": 1,
            "min_block_hours": 1.0,
        },
        "load_linear_daily": {
            "p_min": 10.0,
            "p_max": 30.0,
        },
        "subscription_capacity": {
            "tier_caps_kw": "3.6,7,11",
            "tier_fees": "10,20,30",
            "subscribed_tier_index": 2,
            "penalty_add": 1.5,
        },
    }


def test_create_methods_from_config_builds_all_four_methods():
    methods = create_methods_from_config(_sample_config())

    assert set(methods) == set(METHOD_SHEET_NAMES)
    assert methods["topn_peak_reference_day"].n_low_peaks == 1
    assert methods["topn_peak_reference_day"].window_size_hours_high == 6.0
    assert methods["quantile_daily_budget"].selection_mode == "contiguous"
    assert methods["load_linear_daily"].p_max == 30.0
    assert methods["subscription_capacity"].subscribed_cap_kw == 7.0


def test_create_methods_from_config_rejects_missing_method():
    config = _sample_config()
    del config["load_linear_daily"]
    with pytest.raises(ValueError, match="missing method"):
        create_methods_from_config(config)


def test_create_methods_from_config_rejects_unknown_method():
    config = _sample_config()
    config["unknown_method"] = {"base_fee": 1.0}
    with pytest.raises(ValueError, match="unknown method"):
        create_methods_from_config(config)


def test_create_methods_from_config_propagates_missing_required_param():
    config = _sample_config()
    del config["load_linear_daily"]["p_min"]
    with pytest.raises(ValueError, match="load_linear_daily"):
        create_methods_from_config(config)


def test_load_methods_config_from_example_workbook():
    config = load_methods_config_from_excel(EXAMPLE_CONFIG)

    assert set(config) == set(METHOD_SHEET_NAMES)
    assert config["topn_peak_reference_day"]["base_fee"] == 12.0
    assert config["topn_peak_reference_day"]["use_reference_day"] is True
    assert "time_window_start_hour" not in config["topn_peak_reference_day"]
    assert config["quantile_daily_budget"]["selection_mode"] == "distributed"
    assert config["subscription_capacity"]["tier_caps_kw"] == "3.6,7,11"


def test_load_methods_config_from_excel_missing_sheet(tmp_path):
    path = tmp_path / "incomplete.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "topn_peak_reference_day"
    sheet.append(["parameter", "value", "description"])
    sheet.append(["base_fee", 12.0, "fee"])
    workbook.save(path)

    with pytest.raises(ValueError, match="missing sheet"):
        load_methods_config_from_excel(path)


def test_parse_params_sheet_skips_empty_values(tmp_path):
    path = tmp_path / "config.xlsx"
    workbook = Workbook()
    workbook.remove(workbook.active)

    for sheet_name in METHOD_SHEET_NAMES:
        sheet = workbook.create_sheet(title=sheet_name)
        sheet.append(["parameter", "value", "description"])
        if sheet_name == "topn_peak_reference_day":
            sheet.append(["base_fee", 12.0, "fee"])
            sheet.append(["relative_fee_reduction", 0.4, "reduction"])
            sheet.append(["relative_fee_surcharge", 0.2, "surcharge"])
            sheet.append(["time_window_start_hour", None, "optional"])
        elif sheet_name == "quantile_daily_budget":
            sheet.append(["base_fee", 12.0, "fee"])
            sheet.append(["relative_fee_reduction", 0.4, "reduction"])
            sheet.append(["relative_fee_surcharge", 0.2, "surcharge"])
            sheet.append(["q_low", 0.15, "low"])
            sheet.append(["q_high", 0.15, "high"])
        elif sheet_name == "load_linear_daily":
            sheet.append(["p_min", 10.0, "min"])
            sheet.append(["p_max", 30.0, "max"])
        else:
            sheet.append(["tier_caps_kw", "3.6,7", "caps"])
            sheet.append(["tier_fees", "10,20", "fees"])
            sheet.append(["subscribed_tier_index", 1, "tier"])
            sheet.append(["penalty_add", 1.0, "penalty"])

    workbook.save(path)
    config = load_methods_config_from_excel(path)

    assert "time_window_start_hour" not in config["topn_peak_reference_day"]
    methods = create_methods_from_config(config)
    assert methods["topn_peak_reference_day"].time_window_start_hour is None
